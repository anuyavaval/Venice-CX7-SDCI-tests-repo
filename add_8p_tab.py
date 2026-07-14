#!/usr/bin/env python3
"""
Append an "8-Core Study" tab (SDCI ENABLED) to the existing 1C study workbook.
Reads the 8P results (aggregate + per-core) and adds one tab, leaving the
System and 1-Core Study tabs untouched.

Output (in place): 1P/results/CX7-1C-study-VeniceB0-G2-BIOS.xlsx
"""
import os, csv, glob, sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HERE = os.path.dirname(os.path.abspath(__file__))
WB_PATH = os.path.join(HERE, "1P", "results", "CX7-1C-study-VeniceB0-G2-BIOS.xlsx")
P8_DIR = os.environ.get("P8_DIR") or max(
    glob.glob(os.path.join(HERE, "8P", "results_*")), key=os.path.getmtime)

# 8P layout
CORES = list(range(5, 13))        # 5..12
SIBLINGS = [c + 256 for c in CORES]
NUM_CORES = len(CORES)

AMD_RED="ED1C24"; AMD_GREY="58595B"; HDR_BLUE="1F4E78"; LT_GREY="F2F2F2"
BEST_GRN="C6EFCE"; WHITE="FFFFFF"
thin=Side(style="thin", color="BFBFBF"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)

def title(ws,r,c1,c2,t,bg=AMD_RED,fg=WHITE,sz=13,ht=26):
    ws.merge_cells(start_row=r,start_column=c1,end_row=r,end_column=c2)
    c=ws.cell(r,c1,t); c.font=Font(bold=True,color=fg,size=sz)
    c.fill=PatternFill("solid",fgColor=bg); c.alignment=Alignment(horizontal="left",vertical="center",indent=1)
    ws.row_dimensions[r].height=ht

def load_avgs(d):
    rows={}
    with open(os.path.join(d,"averages.csv")) as f:
        for rec in csv.reader(f):
            if rec and rec[0].isdigit(): rows[int(rec[0])]=rec
    return rows

def load_percore(d):
    path=os.path.join(d,"percore.csv")
    if not os.path.exists(path): return {}
    acc={}
    with open(path) as f:
        rdr=csv.reader(f); next(rdr)
        for rec in rdr:
            if not rec or not rec[0].isdigit(): continue
            P=int(rec[0]); vals=[float(x) for x in rec[2:2+NUM_CORES]]
            acc.setdefault(P,[]).append(vals)
    return {P:[sum(col)/len(col) for col in zip(*lst)] for P,lst in acc.items()}

wb = openpyxl.load_workbook(WB_PATH)
if "8-Core Study" in wb.sheetnames:
    del wb["8-Core Study"]      # rebuild cleanly if re-run
ws2 = wb.create_sheet("8-Core Study")
ws2.sheet_view.showGridLines = False

avgs = load_avgs(P8_DIR); pc = load_percore(P8_DIR)
plist = sorted(avgs.keys())
ncol = 6 + NUM_CORES
for j,w in enumerate([8,15,11,11,9,14]+[9]*NUM_CORES, start=1):
    ws2.column_dimensions[openpyxl.utils.get_column_letter(j)].width=w

title(ws2,1,1,ncol,"CX7 8-Core iPerf Core-Scaling — SDCI ENABLED",bg=AMD_RED,sz=14,ht=30)
title(ws2,2,1,ncol,
      f"eth2 (CX7 400G) | 8Q | iperf cores {CORES[0]}-{CORES[-1]}, 64 IRQs interleaved to "
      f"siblings {SIBLINGS[0]}-{SIBLINGS[-1]} (i%%{NUM_CORES}) | 60s x 10 iters | BIOS G2 | Venice B0",
      bg=AMD_GREY,sz=9,ht=20)
title(ws2,4,1,ncol,"8-CORE STUDY — Aggregate throughput + per-core breakdown (SDCI ENABLED)",bg=HDR_BLUE,sz=11)

hr=5
heads=["-P","Agg (Gbps)","Min","Max","CV %","Retrans"]+[f"core{c}" for c in CORES]
for j,h in enumerate(heads,start=1):
    c=ws2.cell(hr,j,h); c.font=Font(bold=True,color=WHITE); c.fill=PatternFill("solid",fgColor=HDR_BLUE)
    c.alignment=Alignment(horizontal="center",vertical="center"); c.border=BORDER
ws2.row_dimensions[hr].height=18

best=max(plist,key=lambda P:float(avgs[P][1]))
rr=hr+1
for P in plist:
    a=avgs[P]; row=[P,float(a[1]),float(a[2]),float(a[3]),float(a[4]),float(a[5])]
    row+=[round(x,2) for x in pc.get(P,[0]*NUM_CORES)]
    for j,v in enumerate(row,start=1):
        c=ws2.cell(rr,j,v); c.alignment=Alignment(horizontal="center",vertical="center"); c.border=BORDER
        if j in (2,3,4,5) or j>6: c.number_format="0.00"
        if P==best: c.fill=PatternFill("solid",fgColor=BEST_GRN); c.font=Font(bold=True)
        elif rr%2==0: c.fill=PatternFill("solid",fgColor=LT_GREY)
    ws2.row_dimensions[rr].height=15; rr+=1

# per-core note + observations
rr+=1
title(ws2,rr,1,ncol,"KEY OBSERVATIONS",bg=HDR_BLUE,sz=10); rr+=1
notes=[
    f"Best aggregate: -P {best} -> {float(avgs[best][1]):.2f} Gbps (CV {float(avgs[best][4]):.2f}%).",
    "8 cores plateau ~330 Gbps from -P 8 up — about 3.1x the single-core SDCI-on ~107 Gbps (sub-linear; "
    "aggregate ceiling is the shared NIC/PCIe/400G path, not per-core).",
    "Per-core columns (core5-core12) show load balance — each carries ~40-45 Gbps in the plateau.",
    "Retransmits climb with -P (133K @ P8 -> 676K @ P32); P8-P12 is the stable sweet spot "
    "(full throughput, lower retrans, CV ~1.2%). -P 4 was unstable (CV 12.6%).",
    "Settings identical to 1-Core study except 8Q + 8 procs + interleaved IRQs. SDCI ENABLED.",
]
for n in notes:
    ws2.merge_cells(start_row=rr,start_column=1,end_row=rr,end_column=ncol)
    c=ws2.cell(rr,1,"•  "+n); c.alignment=Alignment(vertical="center",indent=1,wrap_text=True)
    ws2.row_dimensions[rr].height=28; rr+=1

wb.save(WB_PATH)
print("Saved:", WB_PATH)
print("Tabs:", wb.sheetnames)
print(f"8P from {P8_DIR} | best -P {best} -> {avgs[best][1]} Gbps")
