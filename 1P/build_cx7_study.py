#!/usr/bin/env python3
"""
Build CX7-1C-study-VeniceB0-G2-BIOS.xlsx
  Tab 1 "System"        — HW / NIC / BIOS / FW / OS details + iperf3 commands
  Tab 2 "1-Core Study"  — -P sweep results: SDCI-off, SDCI-on, and delta
Output: CX7/results/CX7-1C-study-VeniceB0-G2-BIOS.xlsx
"""
import os, csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HERE = os.path.dirname(os.path.abspath(__file__))
OFF_DIR = os.path.join(HERE, "results_20260710_142856")   # SDCI OFF
ON_DIR  = os.path.join(HERE, "results_20260710_190626")   # SDCI ON
OUTDIR = os.path.join(HERE, "results")
os.makedirs(OUTDIR, exist_ok=True)
OUT = os.path.join(OUTDIR, "CX7-1C-study-VeniceB0-G2-BIOS.xlsx")

def load_avgs(d):
    rows = {}
    with open(os.path.join(d, "averages.csv")) as f:
        for rec in csv.reader(f):
            if rec and rec[0].isdigit():
                rows[int(rec[0])] = rec  # P -> [P,avg,min,max,cv,ret]
    return rows

# ─── styles ──────────────────────────────────────────────────────────────────
AMD_RED   = "ED1C24"
AMD_GREY  = "58595B"
HDR_BLUE  = "1F4E78"
LT_BLUE   = "DDEBF7"
LT_GREY   = "F2F2F2"
BEST_GRN  = "C6EFCE"
WHITE = "FFFFFF"
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def title(ws, r, c1, c2, text, bg=AMD_RED, fg=WHITE, sz=13, ht=26):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    cell = ws.cell(r, c1, text)
    cell.font = Font(bold=True, color=fg, size=sz)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[r].height = ht

def kv(ws, r, key, val, keyw=True):
    a = ws.cell(r, 1, key); b = ws.cell(r, 2, val)
    a.font = Font(bold=True, color="333333"); a.alignment = Alignment(vertical="center", indent=1)
    b.alignment = Alignment(vertical="center", indent=1, wrap_text=True)
    a.fill = PatternFill("solid", fgColor=LT_GREY)
    a.border = BORDER; b.border = BORDER
    ws.row_dimensions[r].height = 16

# ═══════════════════════ WORKBOOK ═══════════════════════
wb = openpyxl.Workbook()

# ────────────── Tab 1: System ──────────────
ws = wb.active
ws.title = "System"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 78

title(ws, 1, 1, 2, "CX7 1-Core iPerf Study — System Configuration", bg=AMD_RED, sz=14, ht=30)
title(ws, 2, 1, 2, "Platform: Venice B0  |  BIOS: G2 (PCOV207040_G2N)  |  SUT: congo-0573-host",
      bg=AMD_GREY, sz=10, ht=18)
r = 4

title(ws, r, 1, 2, "PLATFORM / CPU", bg=HDR_BLUE, sz=11); r += 1
for k, v in [
    ("Platform", "AMD Venice B0"),
    ("System (SUT)", "congo-0573-host  (mgmt 10.230.104.91)"),
    ("Load Generator", "galena-3666-host  (mgmt 10.230.96.142)"),
    ("System Product", "AMD Corporation — Congo"),
    ("CPU Model", "AMD Eng Sample: 100-000001041-03"),
    ("Topology", "1 socket, 256 cores/socket, 2 threads/core = 512 CPUs"),
    ("Stepping", "0 (B0)"),
    ("CPU Freq (forced cclk)", "2500 MHz (core5 scaling_cur_freq = 2500 MHz)"),
    ("CPU max / min MHz", "4008.79 / 1500.00"),
    ("NUMA", "2 nodes — node0: 0-127,256-383 | node1: 128-255,384-511"),
    ("Governor / Boost", "performance / boost OFF"),
]:
    kv(ws, r, k, v); r += 1

r += 1
title(ws, r, 1, 2, "BIOS / FIRMWARE", bg=HDR_BLUE, sz=11); r += 1
for k, v in [
    ("BIOS Vendor", "AMD Corporation"),
    ("BIOS Version", "PCOV207040_G2N   (\"G2\" BIOS)"),
    ("BIOS Release Date", "06/25/2026"),
    ("DF EnSrcDnCnRply", "0x1  (set on all IOM/IOD — verified via ARX)"),
    ("BMC Firmware", "n/a (not exposed to OS)"),
    ("FPGA", "n/a (not exposed to OS)"),
    ("OS", "Ubuntu 24.04.3 LTS"),
    ("Kernel", "6.8.0-117-generic"),
]:
    kv(ws, r, k, v); r += 1

r += 1
title(ws, r, 1, 2, "NIC UNDER TEST — ConnectX-7 (eth2)", bg=HDR_BLUE, sz=11); r += 1
for k, v in [
    ("Device", "NVIDIA/Mellanox ConnectX-7 (MT2910)"),
    ("Part Number", "MCX75310AAS-NEA_Ax"),
    ("Description", "ConnectX-7 HHHL, 400GbE / NDR IB, Single-port OSFP, PCIe 5.0 x16"),
    ("PSID", "MT_0000000838"),
    ("Mellanox FW", "28.48.1000 (MT_0000000838)"),
    ("Driver", "mlx5_core  v26.01-1.0.0"),
    ("PCI BDF", "0000:21:00.0  (mst: /dev/mst/mt4129_pciconf0, mlx5_0)"),
    ("PCIe Link", "Gen5 32GT/s x16 (LnkCap = LnkSta, full width)"),
    ("Netdev / MAC", "eth2 / cc:40:f3:8d:56:24"),
    ("Data IP", "192.168.10.2/24  (peer loadgen enp1s0np0 = 192.168.10.3)"),
    ("Link Speed", "400,000 Mb/s (400G), Full duplex, DAC, Link UP"),
    ("MTU", "1500"),
]:
    kv(ws, r, k, v); r += 1

r += 1
title(ws, r, 1, 2, "NIC SETTINGS (as tested)", bg=HDR_BLUE, sz=11); r += 1
for k, v in [
    ("Queues (combined)", "1  (1Q — forced via 'ethtool -L eth2 combined 1')"),
    ("IRQ affinity", "ALL eth2 IRQs (mlx5_comp*@21:00.0) -> core 261"),
    ("iperf pinning", "iperf3 client pinned to core 5 (taskset -c 5)"),
    ("SMT sibling map", "core 5 <-> 261 (CCD0: sibling = core + 256)"),
    ("Ring size (rx/tx)", "2048 / 2048  (max 8192)"),
    ("Flow control", "rx OFF, tx OFF, autoneg OFF (ethtool -A)"),
    ("CQE_COMPRESSION", "AGGRESSIVE(1)"),
    ("PCI_WR_ORDERING", "force_relax(1)"),
    ("LINK_TYPE", "ETH(2)"),
    ("irqbalance", "stopped"),
]:
    kv(ws, r, k, v); r += 1

r += 1
title(ws, r, 1, 2, "TEST METHOD", bg=HDR_BLUE, sz=11); r += 1
for k, v in [
    ("Tool", "iperf3 (single process, single core; -P = parallel TCP streams)"),
    ("Direction", "SUT client -> LoadGen server (unidirectional TCP Rx at server)"),
    ("Duration / Iterations", "60 s per run, 5 iterations per -P (avg reported)"),
    ("-P sweep", "1, 2, 4, 8, 12, 16, 20, 24, 28, 32"),
    ("SDCI State", "Studied both: SDCI OFF and SDCI ENABLED (identical settings)"),
    ("Metric", "sum_received throughput (Gbps) + TCP retransmits"),
]:
    kv(ws, r, k, v); r += 1

# ── iperf3 commands issued ──
r += 1
title(ws, r, 1, 2, "iPERF3 COMMANDS ISSUED", bg=HDR_BLUE, sz=11); r += 1

def cmd_block(r, label, cmd):
    a = ws.cell(r, 1, label); b = ws.cell(r, 2, cmd)
    a.font = Font(bold=True, color="333333"); a.alignment = Alignment(vertical="center", indent=1)
    a.fill = PatternFill("solid", fgColor=LT_GREY); a.border = BORDER
    b.font = Font(name="Consolas", size=9); b.alignment = Alignment(vertical="center", indent=1, wrap_text=True)
    b.border = BORDER
    ws.row_dimensions[r].height = 30
    return r + 1

r = cmd_block(r, "LoadGen (galena-3666-host)  server",
              "taskset -c 5 iperf3 -s -B 192.168.10.3 -p 5201")
r = cmd_block(r, "  loadgen NIC tuning",
              "systemctl stop irqbalance ; ethtool -A enp1s0np0 rx off tx off ; "
              "cpupower frequency-set -g performance ; echo 0 > /sys/devices/system/cpu/cpufreq/boost")
r = cmd_block(r, "SUT (congo-0573-host)  client",
              "taskset -c 5 iperf3 -c 192.168.10.3 -p 5201 -P <N> -t 60 -J")
r = cmd_block(r, "  where -P <N> =",
              "1, 2, 4, 8, 12, 16, 20, 24, 28, 32   (5 iterations each)")
r = cmd_block(r, "  SUT NIC tuning",
              "systemctl stop irqbalance ; ethtool -A eth2 rx off tx off ; "
              "cpupower frequency-set -g performance ; echo 0 > /sys/devices/system/cpu/cpufreq/boost")
r = cmd_block(r, "  force 1Q combined",
              "ethtool -L eth2 combined 1")
r = cmd_block(r, "  steer all eth2 IRQs -> core 261",
              "BDF=$(ethtool -i eth2 | awk '/bus-info:/{print $2}') ; "
              "for irq in $(grep \"$BDF\" /proc/interrupts | sed -E 's/^ *([0-9]+):.*/\\1/'); "
              "do echo 261 > /proc/irq/$irq/smp_affinity_list ; done")

# ────────────── Tab 2: 1-Core Study ──────────────
ws2 = wb.create_sheet("1-Core Study")
ws2.sheet_view.showGridLines = False
for col, w in zip("ABCDEFG", [10, 16, 12, 12, 10, 16, 14]):
    ws2.column_dimensions[col].width = w

title(ws2, 1, 1, 7, "CX7 1-Core iPerf Core-Scaling — SDCI OFF vs ENABLED", bg=AMD_RED, sz=14, ht=30)
title(ws2, 2, 1, 7,
      "eth2 (ConnectX-7, 400G) | 1Q | iperf core 5, IRQ->261 | 60s x 5 iters | BIOS G2 (PCOV207040_G2N) | Venice B0",
      bg=AMD_GREY, sz=9, ht=20)

off = load_avgs(OFF_DIR)
on  = load_avgs(ON_DIR)
plist = sorted(off.keys())

def draw_table(start_row, label, data, sdci_lbl, color):
    title(ws2, start_row, 1, 7, label, bg=color, sz=11)
    hr = start_row + 1
    heads = ["-P", "Avg (Gbps)", "Min", "Max", "CV %", "Retrans", "SDCI"]
    for j, h in enumerate(heads, start=1):
        c = ws2.cell(hr, j, h)
        c.font = Font(bold=True, color=WHITE); c.fill = PatternFill("solid", fgColor=HDR_BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border = BORDER
    ws2.row_dimensions[hr].height = 18
    best_p = max(data, key=lambda p: float(data[p][1]))
    rr = hr + 1
    for P in plist:
        rec = data[P]
        is_best = (P == best_p)
        vals = [P, float(rec[1]), float(rec[2]), float(rec[3]), float(rec[4]), float(rec[5]), sdci_lbl]
        for j, v in enumerate(vals, start=1):
            c = ws2.cell(rr, j, v)
            c.alignment = Alignment(horizontal="center", vertical="center"); c.border = BORDER
            if j in (2, 3, 4, 5): c.number_format = "0.00"
            if is_best:
                c.fill = PatternFill("solid", fgColor=BEST_GRN); c.font = Font(bold=True)
            elif rr % 2 == 0:
                c.fill = PatternFill("solid", fgColor=LT_GREY)
        ws2.row_dimensions[rr].height = 15
        rr += 1
    return rr, best_p

# ── Table 1: SDCI OFF ──
r, off_best = draw_table(4, "TABLE 1  —  SDCI OFF", off, "OFF", HDR_BLUE)

# ── Table 2: SDCI ENABLED ──
r += 1
r, on_best = draw_table(r, "TABLE 2  —  SDCI ENABLED", on, "ON", AMD_RED)

# ── Table 3: Comparison (delta) ──
r += 1
title(ws2, r, 1, 7, "TABLE 3  —  COMPARISON  (SDCI ON vs OFF)", bg=AMD_GREY, sz=11); r += 1
heads = ["-P", "OFF (Gbps)", "ON (Gbps)", "Delta (Gbps)", "Gain %", "", ""]
for j, h in enumerate(heads, start=1):
    c = ws2.cell(r, j, h)
    c.font = Font(bold=True, color=WHITE); c.fill = PatternFill("solid", fgColor=HDR_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center"); c.border = BORDER
ws2.row_dimensions[r].height = 18
r += 1
GOOD = "C6EFCE"; BAD = "FFC7CE"
for P in plist:
    o = float(off[P][1]); n = float(on[P][1]); d = n - o; g = (d / o * 100) if o else 0
    vals = [P, o, n, d, g, "", ""]
    for j, v in enumerate(vals, start=1):
        c = ws2.cell(r, j, v)
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border = BORDER
        if j in (2, 3, 4): c.number_format = "0.00"
        if j == 5: c.number_format = '+0.0"%";-0.0"%"'
        if j in (4, 5):
            c.fill = PatternFill("solid", fgColor=GOOD if d >= 0 else BAD)
            c.font = Font(bold=True)
    ws2.row_dimensions[r].height = 15
    r += 1

# ── notes ──
r += 1
title(ws2, r, 1, 7, "KEY OBSERVATIONS", bg=HDR_BLUE, sz=10); r += 1
maxgain = max(plist, key=lambda P: float(on[P][1]) - float(off[P][1]))
mg_off = float(off[maxgain][1]); mg_on = float(on[maxgain][1])
notes = [
    f"SDCI OFF peak: -P {off_best} -> {float(off[off_best][1]):.2f} Gbps.  "
    f"SDCI ON peak: -P {on_best} -> {float(on[on_best][1]):.2f} Gbps.",
    "SDCI ENABLED lifts single-core throughput from ~89 Gbps to a sustained ~107 Gbps "
    "(P8-P32) — roughly +20 Gbps / +21%.",
    "SDCI reaches its plateau far earlier: full throughput by -P 4 (110.67 Gbps peak) vs "
    "SDCI-off needing -P 32 to hit its lower 89.26 Gbps peak.",
    "Low-concurrency exceptions: -P 1 is a wash; -P 2 is slightly WORSE with SDCI on (-2.6 Gbps). "
    "SDCI benefit needs >=4 concurrent streams.",
    "Stability strong in both (CV mostly <1%). SDCI-on showed retransmits only at -P 28/32 "
    "(2654 / 234 avg); all other points zero.",
    "Both runs used IDENTICAL settings (1Q, iperf core 5, all eth2 IRQs->261, 60s x 5, cclk 2500, "
    "flow-ctrl off). Only SDCI state differs.",
]
for n in notes:
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    c = ws2.cell(r, 1, "•  " + n)
    c.alignment = Alignment(vertical="center", indent=1, wrap_text=True)
    ws2.row_dimensions[r].height = 30
    r += 1

wb.save(OUT)
print("WROTE:", OUT)
print(f"SDCI OFF best -P {off_best} -> {off[off_best][1]} | SDCI ON best -P {on_best} -> {on[on_best][1]}")
