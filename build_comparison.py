#!/usr/bin/env python3
"""
Build CX7-Core-Scaling-Comparison-VeniceB0-G2-BIOS.xlsx
  Tab 1 "Comparison"  — side-by-side aggregate throughput for 1P/2P/8P/16P
  Tab 2 "Chart Data"  — clean table + embedded line chart
Reads the fixed result dirs for each core-count study.
Output: CX7/CX7-Core-Scaling-Comparison-VeniceB0-G2-BIOS.xlsx
"""
import os, csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference

HERE = os.path.dirname(os.path.abspath(__file__))

AMD_RED  = "ED1C24"; AMD_GREY = "58595B"; HDR_BLUE = "1F4E78"
LT_GREY  = "F2F2F2"; WHITE    = "FFFFFF"; BEST_GRN = "C6EFCE"
COLS = {
    "1P":  "BDD7EE",
    "2P":  "FCE4D6",
    "8P":  "E2EFDA",
    "16P": "FFF2CC",
}
thin  = Side(style="thin",   color="BFBFBF")
thick = Side(style="medium", color="8EA9C1")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def thick_border(left=False, right=False):
    return Border(left=thick if left else thin,
                  right=thick if right else thin,
                  top=thin, bottom=thin)

def title(ws, r, c1, c2, t, bg=AMD_RED, fg=WHITE, sz=13, ht=26):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    c = ws.cell(r, c1, t)
    c.font = Font(bold=True, color=fg, size=sz)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[r].height = ht

def hdr(ws, r, c, t, bg=HDR_BLUE, fg=WHITE, sz=9):
    cell = ws.cell(r, c, t)
    cell.font = Font(bold=True, color=fg, size=sz)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER
    return cell

def val(ws, r, c, v, bg=None, bold=False, fmt="0.00", border=BORDER):
    cell = ws.cell(r, c, v)
    cell.font = Font(bold=bold, color="111111")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = fmt
    cell.border = border
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    return cell

def load_avgs(path):
    rows = {}
    with open(path) as f:
        rdr = csv.reader(f); next(rdr)
        for rec in rdr:
            if rec and rec[0].isdigit():
                rows[int(rec[0])] = rec
    return rows

# ── Load all four studies ──────────────────────────────────────────────────────
REPO = os.path.join(HERE, "..", "..", "..")  # worktree -> real repo root
REPO = os.path.normpath(os.path.join(HERE, "../../.."))

# paths relative to the repo root (CX7/)
CX7 = os.path.normpath(os.path.join(HERE, "../../.."))
studies = {
    "1P":  os.path.join(CX7, "1P",  "results_20260710_142856", "averages.csv"),
    "2P":  os.path.join(CX7, "2P",  "results_20260714_160747", "averages.csv"),
    "8P":  os.path.join(CX7, "8P",  "results_20260713_144947", "averages.csv"),
    "16P": os.path.join(CX7, "16P", "results_20260714_100042", "averages.csv"),
}
data   = {k: load_avgs(v) for k, v in studies.items()}
P_LIST = sorted(set().union(*[d.keys() for d in data.values()]))
LABELS = ["1P", "2P", "8P", "16P"]

wb = openpyxl.Workbook()

# ══════════════════════════════════════════════════════════════════════════════
# Tab 1: Comparison
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.active; ws.title = "Comparison"; ws.sheet_view.showGridLines = False

# col 1 = -P  |  per study: Agg, Min, Max, CV%  (4 cols)  →  cols 2-17
STUDY_START = {"1P": 2, "2P": 6, "8P": 10, "16P": 14}
NCOLS = 17
for j, w in enumerate(
    [6, 12, 9, 9, 7,  12, 9, 9, 7,  12, 9, 9, 7,  12, 9, 9, 7], start=1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = w

title(ws, 1, 1, NCOLS,
      "CX7 Core-Scaling Study — Aggregate iPerf Throughput Comparison  (SDCI ENABLED)",
      bg=AMD_RED, sz=14, ht=32)
title(ws, 2, 1, NCOLS,
      "Platform: Venice B0  |  BIOS: G2 (PCOV207040_G2N)  |  SUT: congo-0573-host  |  "
      "60 s × 10 iters per -P  |  TCP Rx  |  64 IRQs interleaved across SMT siblings",
      bg=AMD_GREY, sz=9, ht=18)

# study header band
r = 4
ws.row_dimensions[r].height = 22
c0 = ws.cell(r, 1, "-P")
c0.font = Font(bold=True, color=WHITE, size=10)
c0.fill = PatternFill("solid", fgColor=HDR_BLUE)
c0.alignment = Alignment(horizontal="center", vertical="center")
c0.border = BORDER

for label, sc in STUDY_START.items():
    nc = label.replace("P", "")
    ws.merge_cells(start_row=r, start_column=sc, end_row=r, end_column=sc+3)
    c = ws.cell(r, sc, f"{label}  ({nc}-core, {nc}Q)")
    c.font = Font(bold=True, color="111111", size=11)
    c.fill = PatternFill("solid", fgColor=COLS[label])
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thick_border(left=True, right=True)

# sub-headers
r = 5; ws.row_dimensions[r].height = 30
hdr(ws, r, 1, "-P", bg=HDR_BLUE)
for label, sc in STUDY_START.items():
    for i, h in enumerate(["Agg\n(Gbps)", "Min\n(Gbps)", "Max\n(Gbps)", "CV %"]):
        hdr(ws, r, sc+i, h, bg=HDR_BLUE if i == 0 else "4F81BD", fg=WHITE, sz=9)

# data rows
rr = 6
for P in P_LIST:
    ws.row_dimensions[rr].height = 15
    row_bg = LT_GREY if rr % 2 == 0 else WHITE
    c = ws.cell(rr, 1, P)
    c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor=row_bg)
    c.alignment = Alignment(horizontal="center", vertical="center"); c.border = BORDER

    for label, sc in STUDY_START.items():
        best_P = max(data[label], key=lambda x: float(data[label][x][1]))
        d = data[label].get(P)
        agg = float(d[1]) if d else None
        vals = [agg, float(d[2]) if d else None,
                float(d[3]) if d else None, float(d[4]) if d else None]
        is_best = (P == best_P)
        for i, v in enumerate(vals):
            bg = BEST_GRN if (is_best and i == 0) else (COLS[label] if rr % 2 == 0 else None)
            val(ws, rr, sc+i, v if v is not None else "—",
                bg=bg, bold=(is_best and i == 0),
                border=thick_border(left=(i == 0), right=(i == 3)))
    rr += 1

# best-P summary
rr += 1
title(ws, rr, 1, NCOLS, "BEST -P PER STUDY  (highlighted green in table above)",
      bg=HDR_BLUE, sz=10, ht=20); rr += 1
ws.row_dimensions[rr].height = 20
ws.cell(rr, 1, "Best").font = Font(bold=True)
ws.cell(rr, 1).alignment = Alignment(horizontal="center", vertical="center")
ws.cell(rr, 1).border = BORDER
for label, sc in STUDY_START.items():
    best_P = max(data[label], key=lambda x: float(data[label][x][1]))
    best_v = float(data[label][best_P][1])
    cv_v   = float(data[label][best_P][4])
    ws.merge_cells(start_row=rr, start_column=sc, end_row=rr, end_column=sc+3)
    c = ws.cell(rr, sc, f"-P {best_P}  →  {best_v:.2f} Gbps  (CV {cv_v:.2f}%)")
    c.font = Font(bold=True, color="111111", size=10)
    c.fill = PatternFill("solid", fgColor=BEST_GRN)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thick_border(left=True, right=True)

rr += 2
title(ws, rr, 1, NCOLS, "KEY OBSERVATIONS", bg=HDR_BLUE, sz=10, ht=20); rr += 1

def best_agg(label):
    return float(data[label][max(data[label], key=lambda x: float(data[label][x][1]))][1])

b1, b2, b8, b16 = best_agg("1P"), best_agg("2P"), best_agg("8P"), best_agg("16P")
notes = [
    f"1P (1 core): peak {b1:.1f} Gbps.   2P (2 cores): peak {b2:.1f} Gbps  "
    f"(+{b2-b1:.1f} Gbps vs 1P, {b2/b1:.1f}× scaling).",
    f"8P (8 cores): peak {b8:.1f} Gbps  (+{b8-b2:.1f} Gbps vs 2P, {b8/b1:.1f}× vs 1P).  "
    f"16P (16 cores): peak {b16:.1f} Gbps  (+{b16-b8:.1f} Gbps vs 8P, {b16/b1:.1f}× vs 1P).",
    "16P approaches CX7 400G line rate (~400 Gbps). Diminishing returns beyond 8 cores "
    "reflect NIC saturation rather than CPU bottleneck.",
    "All studies: SDCI ENABLED, Venice B0, G2 BIOS, eth2 (CX7 400G), "
    "IRQs interleaved across SMT siblings via i % NUM_CORES.",
]
for n in notes:
    ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=NCOLS)
    c = ws.cell(rr, 1, "•  " + n)
    c.alignment = Alignment(vertical="center", indent=1, wrap_text=True)
    ws.row_dimensions[rr].height = 28; rr += 1

# ══════════════════════════════════════════════════════════════════════════════
# Tab 2: Chart Data + embedded line chart
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Chart Data"); ws2.sheet_view.showGridLines = False
title(ws2, 1, 1, 5,
      "Aggregate Throughput (Gbps) by -P and Core Count  —  SDCI ENABLED",
      bg=AMD_RED, sz=13, ht=28)
title(ws2, 2, 1, 5,
      "Venice B0  |  G2 BIOS  |  CX7 400G  |  60 s × 10 iters",
      bg=AMD_GREY, sz=9, ht=16)

for j, w in enumerate([7, 15, 15, 15, 15], start=1):
    ws2.column_dimensions[openpyxl.utils.get_column_letter(j)].width = w

hr2 = 4
for j, h in enumerate(["-P", "1-Core (Gbps)", "2-Core (Gbps)",
                        "8-Core (Gbps)", "16-Core (Gbps)"], start=1):
    hdr(ws2, hr2, j, h, bg=HDR_BLUE, fg=WHITE, sz=10)
ws2.row_dimensions[hr2].height = 18

rr2 = hr2 + 1
for P in P_LIST:
    ws2.row_dimensions[rr2].height = 15
    bg = LT_GREY if rr2 % 2 == 0 else None
    val(ws2, rr2, 1, P, bg=bg, fmt="0")
    for j, label in enumerate(LABELS, start=2):
        d = data[label].get(P)
        val(ws2, rr2, j, float(d[1]) if d else "", bg=bg)
    rr2 += 1

# line chart
chart = LineChart()
chart.title = "CX7 Core-Scaling: Aggregate Throughput vs -P  (SDCI ENABLED, Venice B0)"
chart.style = 10
chart.y_axis.title = "Aggregate Throughput (Gbps)"
chart.x_axis.title = "-P (streams per iperf3 process)"
chart.height = 15; chart.width = 26

cats = Reference(ws2, min_col=1, min_row=hr2+1, max_row=rr2-1)
colors_hex = ["4472C4", "ED7D31", "70AD47", "FFC000"]
for col_idx in range(2, 6):
    series_ref = Reference(ws2, min_col=col_idx, min_row=hr2, max_row=rr2-1)
    chart.add_data(series_ref, titles_from_data=True)
chart.set_categories(cats)
for s, color in zip(chart.series, colors_hex):
    s.graphicalProperties.line.solidFill = color
    s.graphicalProperties.line.width = 20000
    s.marker.symbol = "circle"; s.marker.size = 5
    s.marker.graphicalProperties.solidFill = color
    s.marker.graphicalProperties.line.solidFill = color

ws2.add_chart(chart, f"A{rr2+2}")

OUT = os.path.join(CX7, "CX7-Core-Scaling-Comparison-VeniceB0-G2-BIOS.xlsx")
wb.save(OUT)
print("WROTE:", OUT)
