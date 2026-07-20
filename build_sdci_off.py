#!/usr/bin/env python3
"""
Build CX7-Core-Scaling-SDCI-Comparison-VeniceB0-G2-BIOS.xlsx
  Tab 0 "System"   — platform/NIC/BIOS/commands on SUT & LG
  Tab 1-4 "1P".."16P" — SDCI OFF + ON tables, pastel line chart, peak highlighted
  Tab 5 "Summary"  — best-config table + bar chart for all studies
"""
import os, csv, glob
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.marker import Marker
from openpyxl.chart.data_source import NumDataSource, NumRef

# ── CX7 root ───────────────────────────────────────────────────────────────────
CX7 = os.environ.get("CX7_ROOT") or ""
if not CX7:
    _p = os.path.dirname(os.path.abspath(__file__))
    for _ in range(6):
        if os.path.isdir(os.path.join(_p,"1P")) and os.path.isdir(os.path.join(_p,"16P")):
            break
        _p = os.path.dirname(_p)
    CX7 = _p
CX7 = os.path.normpath(CX7)
print("CX7:", CX7)

# ── colours ────────────────────────────────────────────────────────────────────
AMD_RED="ED1C24"; AMD_GREY="58595B"; HDR_BLUE="1F4E78"
LT_GREY="F2F2F2"; WHITE="FFFFFF"
# pastel
P_BLUE  ="AEC6CF"  # SDCI ON  line
P_ORANGE="FFD8A8"  # SDCI OFF line
P_BLUE_D ="4472C4"  # darker for chart line
P_ORANGE_D="ED7D31"
PEAK_ON_BG ="D9EAF7"  # highlight cell ON peak
PEAK_OFF_BG="FFF0D0"  # highlight cell OFF peak
BEST_GRN="C6EFCE"
ON_HDR ="2E75B6"; OFF_HDR="C55A11"
ON_TINT="DDEEFF"; OFF_TINT="FFF0E0"
TBD_BG ="F2F2F2"

thin=Side(style="thin",color="BFBFBF"); thick=Side(style="medium",color="999999")
BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
def thickB(l=False,r=False):
    return Border(left=thick if l else thin,right=thick if r else thin,top=thin,bottom=thin)

def title(ws,r,c1,c2,t,bg=AMD_RED,fg=WHITE,sz=13,ht=26):
    ws.merge_cells(start_row=r,start_column=c1,end_row=r,end_column=c2)
    c=ws.cell(r,c1,t); c.font=Font(bold=True,color=fg,size=sz)
    c.fill=PatternFill("solid",fgColor=bg)
    c.alignment=Alignment(horizontal="left",vertical="center",indent=1)
    ws.row_dimensions[r].height=ht

def kv(ws,r,k,v,bg=LT_GREY):
    a=ws.cell(r,1,k); b=ws.cell(r,2,v)
    a.font=Font(bold=True,color="333333"); a.alignment=Alignment(vertical="center",indent=1)
    b.alignment=Alignment(vertical="center",indent=1,wrap_text=True)
    a.fill=PatternFill("solid",fgColor=bg); a.border=BORDER; b.border=BORDER
    ws.row_dimensions[r].height=16

def hdr(ws,r,c,t,bg=HDR_BLUE,fg=WHITE,sz=9):
    cell=ws.cell(r,c,t); cell.font=Font(bold=True,color=fg,size=sz)
    cell.fill=PatternFill("solid",fgColor=bg)
    cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    cell.border=BORDER; return cell

def val(ws,r,c,v,bg=None,bold=False,fmt="0.00",border=BORDER,italic=False,color="111111"):
    cell=ws.cell(r,c,v); cell.font=Font(bold=bold,italic=italic,color=color)
    cell.alignment=Alignment(horizontal="center",vertical="center")
    cell.number_format=fmt; cell.border=border
    if bg: cell.fill=PatternFill("solid",fgColor=bg)
    return cell

def cmd_row(ws,r,label,cmd_text,NCOLS):
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2)
    a=ws.cell(r,1,label); a.font=Font(bold=True,color="333333")
    a.alignment=Alignment(vertical="center",indent=1)
    a.fill=PatternFill("solid",fgColor=LT_GREY); a.border=BORDER
    ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=NCOLS)
    b=ws.cell(r,3,cmd_text)
    b.font=Font(name="Consolas",size=8,color="1F3864")
    b.alignment=Alignment(vertical="center",indent=1,wrap_text=True); b.border=BORDER
    ws.row_dimensions[r].height=28; return r+1

def load_avgs(path):
    rows={}
    if not path or not os.path.exists(path): return rows
    with open(path) as f:
        rdr=csv.reader(f); next(rdr)
        for rec in rdr:
            if rec and rec[0].isdigit(): rows[int(rec[0])]=rec
    return rows

def best_of(d):
    if not d: return None,None,None
    bp=max(d,key=lambda x:float(d[x][1]))
    return bp,float(d[bp][1]),float(d[bp][4])

# ── data ───────────────────────────────────────────────────────────────────────
OFF_DIRS={
    "1P": os.path.join(CX7,"1P","results_20260716_202543"),
    "2P": os.path.join(CX7,"2P","results_20260716_221130"),
    "8P_a":os.path.join(CX7,"8P","results_20260717_000053"),
    "8P_b":os.path.join(CX7,"8P","results_20260717_063547"),
    "16P":os.path.join(CX7,"16P","results_20260717_073011"),
}
ON_DIRS={
    "1P": os.path.join(CX7,"1P","results_20260717_104422"),
    "2P": os.path.join(CX7,"2P","results_20260717_122930"),
    "8P": os.path.join(CX7,"8P","results_20260717_163411"),
    "16P":os.path.join(CX7,"16P","results_20260717_182032"),
}

def load_off(label):
    if label=="8P":
        d=load_avgs(os.path.join(OFF_DIRS["8P_a"],"averages.csv"))
        da={P:d[P] for P in [1,2,4,8,12] if P in d}
        db=load_avgs(os.path.join(OFF_DIRS["8P_b"],"averages.csv"))
        da.update(db); return da
    return load_avgs(os.path.join(OFF_DIRS[label],"averages.csv"))

def load_on(label):
    d=ON_DIRS.get(label)
    path=os.path.join(d,"averages.csv") if d else None
    if not path or not os.path.exists(path):
        print(f"  {label} ON: not available"); return {}
    print(f"  {label} ON: {os.path.basename(d)}")
    return load_avgs(path)

P_LIST=[1,2,4,8,12,16,20,24,28,32]
STUDIES=[("1P",1),("2P",2),("8P",8),("16P",16)]
data_off={l:load_off(l) for l,_ in STUDIES}
data_on ={l:load_on(l)  for l,_ in STUDIES}

wb=openpyxl.Workbook()

# ══════════════════════════════════════════════════════════════════════════════
# Tab 0: System + Commands
# ══════════════════════════════════════════════════════════════════════════════
ws=wb.active; ws.title="System"; ws.sheet_view.showGridLines=False
ws.column_dimensions["A"].width=28; ws.column_dimensions["B"].width=40
ws.column_dimensions["C"].width=55; ws.column_dimensions["D"].width=14
NCOLS_SYS=4

title(ws,1,1,NCOLS_SYS,"CX7 Core-Scaling iPerf Study — System Configuration  (SDCI OFF vs ON)",bg=AMD_RED,sz=14,ht=30)
title(ws,2,1,NCOLS_SYS,"Platform: Venice B0  |  BIOS: G2 (PCOV207040_G2N)  |  SUT: congo-0573-host",bg=AMD_GREY,sz=10,ht=18)

r=4
title(ws,r,1,NCOLS_SYS,"PLATFORM / CPU",bg=HDR_BLUE,sz=11); r+=1
for k,v_ in [("Platform","AMD Venice B0"),("System (SUT)","congo-0573-host"),
    ("Load Generator","galena-3666-host"),("CPU Model","AMD Eng Sample: 100-000001041-03"),
    ("Topology","1 socket, 256 cores/socket, 2 threads/core = 512 CPUs"),
    ("CPU Freq (cclk)","2500 MHz  (policy max 2.50 GHz, asserted by HW — confirmed via cpupower)"),
    ("Governor / Boost","performance / OFF"),
    ("NUMA","node0: 0-127, 256-383  |  node1: 128-255, 384-511")]:
    kv(ws,r,k,v_); r+=1
r+=1
title(ws,r,1,NCOLS_SYS,"BIOS / FIRMWARE",bg=HDR_BLUE,sz=11); r+=1
for k,v_ in [("BIOS Version","PCOV207040_G2N  (G2 BIOS)"),("BIOS Release Date","06/25/2026"),
    ("DF EnSrcDnCnRply","0x1"),("SDCI","Both states tested — see per-study tabs"),
    ("OS / Kernel","Ubuntu 24.04.3 LTS / 6.8.0-117-generic")]:
    kv(ws,r,k,v_); r+=1
r+=1
title(ws,r,1,NCOLS_SYS,"NIC UNDER TEST — ConnectX-7 (eth2)",bg=HDR_BLUE,sz=11); r+=1
for k,v_ in [("Device","NVIDIA/Mellanox ConnectX-7 (MT2910)"),("Part Number","MCX75310AAS-NEA_Ax"),
    ("Mellanox FW","28.48.1000 (MT_0000000838)"),("Driver","mlx5_core v26.01-1.0.0"),
    ("PCI BDF","0000:21:00.0"),("PCIe Link","Gen5 32GT/s x16"),
    ("Data IP (SUT)","192.168.10.2/24"),("Data IP (LG)","192.168.10.3/24"),
    ("Link Speed","400G, Full duplex, DAC, Link UP"),("MTU","1500")]:
    kv(ws,r,k,v_); r+=1
r+=1
title(ws,r,1,NCOLS_SYS,"TEST METHOD",bg=HDR_BLUE,sz=11); r+=1
for k,v_ in [
    ("Tool","iperf3 — N concurrent processes, one per core; -P parallel streams per process"),
    ("Direction","SUT clients -> LoadGen servers (unidirectional TCP Rx)"),
    ("Duration / Iterations","60 s per run, 10 iterations per -P (aggregate avg reported)"),
    ("-P sweep","1, 2, 4, 8, 12, 16, 20, 24, 28, 32"),
    ("IRQ — 1P","All 64 IRQs -> sibling 261 (SMT sibling of core 5)"),
    ("IRQ — 2P/8P/16P","64 IRQs interleaved across SMT siblings via: irq_index % NUM_CORES"),
    ("Throughput","Aggregate = sum of all proc throughputs; per-core detail in raw JSON")]:
    kv(ws,r,k,v_); r+=1
r+=1

title(ws,r,1,NCOLS_SYS,"COMMANDS ISSUED — LOADGEN (galena-3666-host)",bg=OFF_HDR,sz=11); r+=1
r=cmd_row(ws,r,"Stop irqbalance","sudo systemctl stop irqbalance",NCOLS_SYS)
r=cmd_row(ws,r,"Flow ctrl OFF","sudo ethtool -A enp1s0np0 rx off tx off",NCOLS_SYS)
r=cmd_row(ws,r,"Governor perf","sudo cpupower frequency-set -g performance",NCOLS_SYS)
r=cmd_row(ws,r,"Boost OFF","sudo bash -c 'echo 0 > /sys/devices/system/cpu/cpufreq/boost'",NCOLS_SYS)
r=cmd_row(ws,r,"Kill old servers","sudo pkill -f 'iperf3 -s'",NCOLS_SYS)
r=cmd_row(ws,r,"Start N servers (e.g. 2P)","sudo nohup taskset -c <core_i> iperf3 -s -B 192.168.10.3 -p <5201+i> > /tmp/iperf3_srv_<port>.log 2>&1 &",NCOLS_SYS)
r+=1

title(ws,r,1,NCOLS_SYS,"COMMANDS ISSUED — SUT (congo-0573-host)",bg=ON_HDR,sz=11); r+=1
r=cmd_row(ws,r,"Stop irqbalance","sudo systemctl stop irqbalance",NCOLS_SYS)
r=cmd_row(ws,r,"Flow ctrl OFF","sudo ethtool -A eth2 rx off tx off",NCOLS_SYS)
r=cmd_row(ws,r,"Governor perf","sudo cpupower frequency-set -g performance",NCOLS_SYS)
r=cmd_row(ws,r,"Boost OFF","sudo bash -c 'echo 0 > /sys/devices/system/cpu/cpufreq/boost'",NCOLS_SYS)
r=cmd_row(ws,r,"Set N queues","sudo ethtool -L eth2 combined <N>  # N = 1/2/8/16",NCOLS_SYS)
r=cmd_row(ws,r,"Get NIC BDF","BDF=$(ethtool -i eth2 | awk '/bus-info/{print $2}')",NCOLS_SYS)
r=cmd_row(ws,r,"Steer IRQs (1P)","for irq in $(grep $BDF /proc/interrupts | awk -F: '{print $1}' | tr -d ' '); do echo 261 > /proc/irq/$irq/smp_affinity_list; done",NCOLS_SYS)
r=cmd_row(ws,r,"Interleave IRQs (multi)","i=0; for irq in $(grep $BDF /proc/interrupts | awk -F: '{print $1}' | tr -d ' '); do echo ${SIBLINGS[$((i%N))]} > /proc/irq/$irq/smp_affinity_list; i=$((i+1)); done",NCOLS_SYS)
r=cmd_row(ws,r,"Run iperf3 (1P)","taskset -c 5 iperf3 -c 192.168.10.3 -p 5201 -P <N> -t 60 -J",NCOLS_SYS)
r=cmd_row(ws,r,"Run iperf3 (multi, concurrent)","for i in 0..N-1: taskset -c <core_i> iperf3 -c 192.168.10.3 -p <5201+i> -P <P> -t 60 -J  (run all in parallel via ThreadPoolExecutor)",NCOLS_SYS)
r+=1

title(ws,r,1,NCOLS_SYS,"STUDY CONFIGURATION SUMMARY",bg=HDR_BLUE,sz=11); r+=1
config_rows=[
    ("1P (1-core)","Core 5  |  1Q  |  all 64 IRQs -> sibling 261  |  10 iters x 60s  |  SDCI OFF + ON"),
    ("2P (2-core)","Cores 1-2  |  2Q  |  64 IRQs interleaved -> siblings 257-258 (32/sib)  |  10 iters x 60s"),
    ("8P (8-core)","Cores 5-12  |  8Q  |  64 IRQs interleaved -> siblings 261-268 (8/sib)   |  10 iters x 60s"),
    ("16P (16-core)","Cores 1-16  |  16Q  |  64 IRQs interleaved -> siblings 257-272 (4/sib) |  10 iters x 60s"),
]
for k,v_ in config_rows:
    kv(ws,r,k,v_); r+=1

# ══════════════════════════════════════════════════════════════════════════════
# Per-study tabs
# ══════════════════════════════════════════════════════════════════════════════
def make_tab(wb, label, nc, doff, don):
    ws=wb.create_sheet(label); ws.sheet_view.showGridLines=False
    NCOLS=6
    col_w=[8,15,11,11,9,14]
    for j,w in enumerate(col_w,start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width=w

    # best -P for each (used for highlight & chart marker)
    bp_off,bv_off,cv_off=best_of(doff)
    bp_on, bv_on, cv_on =best_of(don)
    # Use the single best -P from SDCI ON as the common highlight point
    # (as per: "best -P per study")
    # Both OFF and ON will be highlighted at their OWN best -P

    if nc==1: irq_str="all 64 IRQs -> sibling 261"; core_str="core 5"
    else: irq_str=f"64 IRQs interleaved -> siblings {257}..{256+nc} (i%{nc})"; core_str=f"cores 1-{nc}"

    title(ws,1,1,NCOLS,f"CX7 {label} ({nc}-Core) — SDCI OFF vs ON  |  Venice B0, G2 BIOS",bg=AMD_RED,sz=14,ht=30)
    title(ws,2,1,NCOLS,f"eth2 (CX7 400G)  |  {nc}Q  |  {core_str}  |  {irq_str}  |  60s x 10 iters",bg=AMD_GREY,sz=9,ht=20)

    def write_table(ws, start_row, data, sdci_label, hdr_color, tint, peak_bg, bp):
        r=start_row
        title(ws,r,1,NCOLS,f"SDCI {sdci_label} — Aggregate Throughput (Gbps)",bg=hdr_color,sz=11,ht=20); r+=1
        hr=r
        for j,h in enumerate(["-P","Agg avg\n(Gbps)","Min\n(Gbps)","Max\n(Gbps)","CV %","Retrans\navg"],start=1):
            hdr(ws,hr,j,h,bg=hdr_color,sz=9)
        ws.row_dimensions[hr].height=30; r+=1

        for P in P_LIST:
            ws.row_dimensions[r].height=15
            d=data.get(P); is_peak=(P==bp)
            row_bg=tint if (r%2==0 and not is_peak) else (peak_bg if is_peak else None)
            val(ws,r,1,P,bg=row_bg,bold=is_peak,fmt="0",
                border=thickB(l=True) if is_peak else BORDER)
            if d:
                val(ws,r,2,float(d[1]),bg=row_bg,bold=is_peak,
                    border=BORDER);
                val(ws,r,3,float(d[2]),bg=row_bg,border=BORDER)
                val(ws,r,4,float(d[3]),bg=row_bg,border=BORDER)
                val(ws,r,5,float(d[4]),bg=row_bg,border=BORDER)
                val(ws,r,6,float(d[5]),bg=row_bg,
                    border=thickB(r=True) if is_peak else BORDER)
            else:
                for j in range(2,7): val(ws,r,j,"—",bg=TBD_BG,fmt="@")
            r+=1

        # peak banner
        r+=1
        if bp:
            bp_v=float(data[bp][1]); bp_cv=float(data[bp][4])
            ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=NCOLS)
            c=ws.cell(r,1,f"Peak: -P {bp}  ->  {bp_v:.2f} Gbps  (CV {bp_cv:.2f}%)")
            c.font=Font(bold=True,color="111111",size=10)
            c.fill=PatternFill("solid",fgColor=peak_bg)
            c.alignment=Alignment(horizontal="center",vertical="center"); c.border=BORDER
            ws.row_dimensions[r].height=20; r+=1
        return r+1

    r=4
    r=write_table(ws,r,doff,"DISABLED",OFF_HDR,OFF_TINT,PEAK_OFF_BG,bp_off)
    r=write_table(ws,r,don, "ENABLED", ON_HDR, ON_TINT, PEAK_ON_BG, bp_on)

    # SDCI impact
    title(ws,r,1,NCOLS,"SDCI IMPACT",bg=HDR_BLUE,sz=10,ht=18); r+=1
    if doff and don:
        delta=bv_on-bv_off; pct=delta/bv_off*100 if bv_off else 0
        notes=[
            f"SDCI OFF peak: -P {bp_off} -> {bv_off:.2f} Gbps  |  SDCI ON peak: -P {bp_on} -> {bv_on:.2f} Gbps",
            f"Delta: {delta:+.2f} Gbps  ({pct:+.1f}%)  at {nc} core(s)."
        ]
    else:
        notes=["SDCI ON data not yet available."]
    for n in notes:
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=NCOLS)
        c=ws.cell(r,1,"   "+n); c.alignment=Alignment(vertical="center",indent=1,wrap_text=True)
        ws.row_dimensions[r].height=20; r+=1
    r+=1

    # ── Chart data block (hidden rows used as chart source) ──────────────────
    # Place chart data starting at col 8 (off-screen)
    CD=8   # chart data start col
    ws.column_dimensions[openpyxl.utils.get_column_letter(CD)].width=6
    ws.column_dimensions[openpyxl.utils.get_column_letter(CD+1)].width=13
    ws.column_dimensions[openpyxl.utils.get_column_letter(CD+2)].width=13

    cd_r=4   # chart data start row
    ws.cell(cd_r,CD,"-P"); ws.cell(cd_r,CD+1,"SDCI OFF"); ws.cell(cd_r,CD+2,"SDCI ON")
    for i,P in enumerate(P_LIST):
        rr=cd_r+1+i
        ws.cell(rr,CD,P)
        d=doff.get(P); ws.cell(rr,CD+1,float(d[1]) if d else None)
        d=don.get(P);  ws.cell(rr,CD+2,float(d[1]) if d else None)
    cd_end=cd_r+len(P_LIST)

    # ── Line chart ────────────────────────────────────────────────────────────
    chart=LineChart()
    chart.title=f"CX7 {label} — SDCI OFF vs ON  (Aggregate Gbps)"
    chart.style=2
    chart.y_axis.title="Aggregate Throughput (Gbps)"; chart.x_axis.title="-P"
    chart.height=13; chart.width=22
    chart.legend.position="b"

    cats=Reference(ws,min_col=CD,min_row=cd_r+1,max_row=cd_end)

    # OFF series — pastel orange
    s_off=Reference(ws,min_col=CD+1,min_row=cd_r,max_row=cd_end)
    chart.add_data(s_off,titles_from_data=True)
    s=chart.series[0]
    s.graphicalProperties.line.solidFill=P_ORANGE_D
    s.graphicalProperties.line.width=18000
    s.marker.symbol="circle"; s.marker.size=4
    s.marker.graphicalProperties.solidFill=P_ORANGE_D
    s.marker.graphicalProperties.line.solidFill=P_ORANGE_D

    # ON series — pastel blue
    s_on=Reference(ws,min_col=CD+2,min_row=cd_r,max_row=cd_end)
    chart.add_data(s_on,titles_from_data=True)
    s=chart.series[1]
    s.graphicalProperties.line.solidFill=P_BLUE_D
    s.graphicalProperties.line.width=18000
    s.marker.symbol="circle"; s.marker.size=4
    s.marker.graphicalProperties.solidFill=P_BLUE_D
    s.marker.graphicalProperties.line.solidFill=P_BLUE_D

    chart.set_categories(cats)

    # place chart to the right of data tables
    chart_anchor=f"H{r+1}"
    ws.add_chart(chart,chart_anchor)

    # ── Peak annotation in chart data ─────────────────────────────────────────
    # Mark peak rows with bold + background in chart data block
    for i,P in enumerate(P_LIST):
        rr=cd_r+1+i
        if P==bp_off:
            ws.cell(rr,CD+1).fill=PatternFill("solid",fgColor=PEAK_OFF_BG)
            ws.cell(rr,CD+1).font=Font(bold=True)
        if P==bp_on:
            ws.cell(rr,CD+2).fill=PatternFill("solid",fgColor=PEAK_ON_BG)
            ws.cell(rr,CD+2).font=Font(bold=True)

for label,nc in STUDIES:
    make_tab(wb,label,nc,data_off[label],data_on[label])

# ══════════════════════════════════════════════════════════════════════════════
# Tab 5: Summary — best config per study, OFF vs ON + bar chart
# ══════════════════════════════════════════════════════════════════════════════
ws5=wb.create_sheet("Summary"); ws5.sheet_view.showGridLines=False
for j,w in enumerate([10,16,18,16,18,14],start=1):
    ws5.column_dimensions[openpyxl.utils.get_column_letter(j)].width=w

NCOLS5=6
title(ws5,1,1,NCOLS5,"CX7 Core-Scaling — Best Config Summary  |  SDCI OFF vs ON",bg=AMD_RED,sz=14,ht=30)
title(ws5,2,1,NCOLS5,"Venice B0, G2 BIOS  |  Best -P per study per SDCI state  |  Peak aggregate throughput",bg=AMD_GREY,sz=9,ht=18)

r=4
title(ws5,r,1,NCOLS5,"PEAK THROUGHPUT — SDCI OFF vs ON (best -P for each)",bg=HDR_BLUE,sz=11,ht=20); r+=1

# header
hr=r; ws5.row_dimensions[hr].height=32
for j,h in enumerate(["Study","SDCI OFF\nPeak (Gbps)","SDCI OFF\nBest -P","SDCI ON\nPeak (Gbps)","SDCI ON\nBest -P","Delta\n(ON - OFF)"],start=1):
    c=ws5.cell(hr,j,h); c.font=Font(bold=True,color=WHITE,size=10)
    c.fill=PatternFill("solid",fgColor=HDR_BLUE)
    c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    c.border=BORDER
r+=1

for label,nc in STUDIES:
    bp_off,bv_off,_=best_of(data_off[label])
    bp_on, bv_on, _=best_of(data_on[label])
    delta=(bv_on-bv_off) if (bv_off and bv_on) else None
    row_bg=LT_GREY if r%2==0 else None

    val(ws5,r,1,f"{label} ({nc}-core)",bg=row_bg,fmt="@",bold=True)
    val(ws5,r,2,bv_off if bv_off else "—",bg=PEAK_OFF_BG,bold=True)
    val(ws5,r,3,f"-P {bp_off}" if bp_off else "—",bg=PEAK_OFF_BG,fmt="@")
    val(ws5,r,4,bv_on  if bv_on  else "—",bg=PEAK_ON_BG, bold=True)
    val(ws5,r,5,f"-P {bp_on}"  if bp_on  else "—",bg=PEAK_ON_BG, fmt="@")
    if delta is not None:
        dcolor="006100" if delta>=0 else "9C0006"
        dbg  ="C6EFCE"  if delta>=0 else "FFC7CE"
        c=ws5.cell(r,6,f"{delta:+.2f}")
        c.font=Font(bold=True,color=dcolor); c.number_format="0.00"
        c.alignment=Alignment(horizontal="center",vertical="center")
        c.fill=PatternFill("solid",fgColor=dbg); c.border=BORDER
    else:
        val(ws5,r,6,"—",bg=TBD_BG,fmt="@")
    ws5.row_dimensions[r].height=18; r+=1

r+=2
title(ws5,r,1,NCOLS5,"KEY OBSERVATIONS",bg=HDR_BLUE,sz=10,ht=18); r+=1
notes=[
    "1P: SDCI ON gives +1.2 Gbps (+1.1%) — most visible SDCI benefit at single core.",
    "2P: SDCI OFF and ON essentially tied (~101 Gbps); NIC is the bottleneck at 2 cores.",
    "8P: SDCI ON slightly lower peak (-1.2 Gbps); high retransmit counts at 8 cores mask SDCI benefit.",
    "16P: Both states hit near line rate (~377 Gbps). SDCI ON marginally higher (+0.4 Gbps) at -P 12.",
    "Throughput plateau: 2P and 1P converge at ~100-105 Gbps (single-core NIC limit); "
    "scaling resumes at 8P and 16P via multiple queues.",
]
for n in notes:
    ws5.merge_cells(start_row=r,start_column=1,end_row=r,end_column=NCOLS5)
    c=ws5.cell(r,1,"•  "+n); c.alignment=Alignment(vertical="center",indent=1,wrap_text=True)
    ws5.row_dimensions[r].height=26; r+=1

# ── Bar chart: peak OFF vs ON per study ──────────────────────────────────────
r+=1
# chart data block at col 8
CD=8
ws5.cell(r,CD,"Study"); ws5.cell(r,CD+1,"SDCI OFF"); ws5.cell(r,CD+2,"SDCI ON")
cd_start=r
for i,(label,nc) in enumerate(STUDIES):
    rr=r+1+i
    _,bv_off,_=best_of(data_off[label]); _,bv_on,_=best_of(data_on[label])
    ws5.cell(rr,CD,label)
    ws5.cell(rr,CD+1,bv_off if bv_off else 0)
    ws5.cell(rr,CD+2,bv_on  if bv_on  else 0)
cd_end=r+len(STUDIES)

chart=BarChart(); chart.type="col"; chart.grouping="clustered"
chart.title="Peak Throughput: SDCI OFF vs ON"; chart.style=2
chart.y_axis.title="Gbps"; chart.x_axis.title="Core Config"
chart.height=12; chart.width=18; chart.legend.position="b"

cats=Reference(ws5,min_col=CD,min_row=cd_start+1,max_row=cd_end)
s_off=Reference(ws5,min_col=CD+1,min_row=cd_start,max_row=cd_end)
s_on =Reference(ws5,min_col=CD+2,min_row=cd_start,max_row=cd_end)
chart.add_data(s_off,titles_from_data=True)
chart.add_data(s_on, titles_from_data=True)
chart.set_categories(cats)
chart.series[0].graphicalProperties.solidFill=P_ORANGE_D
chart.series[0].graphicalProperties.line.solidFill=P_ORANGE_D
chart.series[1].graphicalProperties.solidFill=P_BLUE_D
chart.series[1].graphicalProperties.line.solidFill=P_BLUE_D
ws5.add_chart(chart,"A"+str(r+1))

OUT=os.path.join(CX7,"CX7-Core-Scaling-SDCI-Comparison-VeniceB0-G2-BIOS.xlsx")
wb.save(OUT)
print("WROTE:",OUT)
for label,_ in STUDIES:
    bp_off,bv_off,_=best_of(data_off[label])
    bp_on, bv_on, _=best_of(data_on[label])
    delta=(bv_on-bv_off) if bv_off and bv_on else 0
    print(f"  {label}: OFF={bv_off:.1f}G(-P{bp_off})  ON={bv_on:.1f}G(-P{bp_on})  delta={delta:+.2f}G")
