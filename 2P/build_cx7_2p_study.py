#!/usr/bin/env python3
"""
Build CX7-2C-study-VeniceB0-G2-BIOS.xlsx  (SDCI ENABLED)
  Tab 1 "System"      — HW/NIC/BIOS/FW + 2P settings + iperf3 commands
  Tab 2 "2-Core Study"— aggregate -P sweep table (SDCI ENABLED) + per-core breakdown
Reads the newest results_* dir (or RESULT_DIR env override).
Output: 2P/results/CX7-2C-study-VeniceB0-G2-BIOS.xlsx
"""
import os, csv, glob
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from cx7_config import CORES, SIBLINGS, NUM_CORES, BASE_PORT

HERE = os.path.dirname(os.path.abspath(__file__))
RESULT_DIR = os.environ.get("RESULT_DIR") or max(
    glob.glob(os.path.join(HERE, "results_*")), key=os.path.getmtime, default="")
OUTDIR = os.path.join(HERE, "results"); os.makedirs(OUTDIR, exist_ok=True)
OUT = os.path.join(OUTDIR, "CX7-2C-study-VeniceB0-G2-BIOS.xlsx")

AMD_RED="ED1C24"; AMD_GREY="58595B"; HDR_BLUE="1F4E78"; LT_GREY="F2F2F2"
BEST_GRN="C6EFCE"; WHITE="FFFFFF"
thin=Side(style="thin", color="BFBFBF"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)

def title(ws,r,c1,c2,t,bg=AMD_RED,fg=WHITE,sz=13,ht=26):
    ws.merge_cells(start_row=r,start_column=c1,end_row=r,end_column=c2)
    c=ws.cell(r,c1,t); c.font=Font(bold=True,color=fg,size=sz)
    c.fill=PatternFill("solid",fgColor=bg); c.alignment=Alignment(horizontal="left",vertical="center",indent=1)
    ws.row_dimensions[r].height=ht

def kv(ws,r,k,v):
    a=ws.cell(r,1,k); b=ws.cell(r,2,v)
    a.font=Font(bold=True,color="333333"); a.alignment=Alignment(vertical="center",indent=1)
    b.alignment=Alignment(vertical="center",indent=1,wrap_text=True)
    a.fill=PatternFill("solid",fgColor=LT_GREY); a.border=BORDER; b.border=BORDER
    ws.row_dimensions[r].height=16

def load_avgs(d):
    rows={}
    with open(os.path.join(d,"averages.csv")) as f:
        for rec in csv.reader(f):
            if rec and rec[0].isdigit(): rows[int(rec[0])]=rec
    return rows

def load_percore(d):
    """Return {P: [avg_core_gbps per core]} averaged over iterations."""
    path=os.path.join(d,"percore.csv")
    if not os.path.exists(path): return {}
    acc={}
    with open(path) as f:
        rdr=csv.reader(f); hdr=next(rdr)
        for rec in rdr:
            if not rec or not rec[0].isdigit(): continue
            P=int(rec[0]); vals=[float(x) for x in rec[2:2+NUM_CORES]]
            acc.setdefault(P,[]).append(vals)
    out={}
    for P,lst in acc.items():
        out[P]=[sum(col)/len(col) for col in zip(*lst)]
    return out

wb=openpyxl.Workbook()

# ── Tab 1: System ──
ws=wb.active; ws.title="System"; ws.sheet_view.showGridLines=False
ws.column_dimensions["A"].width=30; ws.column_dimensions["B"].width=82
title(ws,1,1,2,"CX7 2-Core iPerf Study — System Configuration",bg=AMD_RED,sz=14,ht=30)
title(ws,2,1,2,"Platform: Venice B0 | BIOS: G2 (PCOV207040_G2N) | SUT: congo-0573-host | SDCI: ENABLED",
      bg=AMD_GREY,sz=10,ht=18)
r=4
title(ws,r,1,2,"PLATFORM / CPU",bg=HDR_BLUE,sz=11); r+=1
for k,v in [("Platform","AMD Venice B0"),("System (SUT)","congo-0573-host"),
    ("Load Generator","galena-3666-host"),("CPU Model","AMD Eng Sample: 100-000001041-03"),
    ("Topology","1 socket, 256 cores/socket, 2 threads/core = 512 CPUs"),
    ("CPU Freq (forced cclk)","2500 MHz"),("Governor / Boost","performance / OFF"),
    ("NUMA","node0: 0-127,256-383 | node1: 128-255,384-511")]:
    kv(ws,r,k,v); r+=1
r+=1
title(ws,r,1,2,"BIOS / FIRMWARE",bg=HDR_BLUE,sz=11); r+=1
for k,v in [("BIOS Version","PCOV207040_G2N  (\"G2\" BIOS)"),("BIOS Release Date","06/25/2026"),
    ("DF EnSrcDnCnRply","0x1"),("SDCI State","ENABLED"),
    ("OS / Kernel","Ubuntu 24.04.3 LTS / 6.8.0-117-generic")]:
    kv(ws,r,k,v); r+=1
r+=1
title(ws,r,1,2,"NIC UNDER TEST — ConnectX-7 (eth2)",bg=HDR_BLUE,sz=11); r+=1
for k,v in [("Device","NVIDIA/Mellanox ConnectX-7 (MT2910)"),("Part Number","MCX75310AAS-NEA_Ax"),
    ("Mellanox FW","28.48.1000 (MT_0000000838)"),("Driver","mlx5_core v26.01-1.0.0"),
    ("PCI BDF","0000:21:00.0"),("PCIe Link","Gen5 32GT/s x16"),
    ("Data IP","192.168.10.2/24 (peer loadgen 192.168.10.3)"),
    ("Link Speed","400G, Full duplex, DAC, Link UP"),("MTU","1500")]:
    kv(ws,r,k,v); r+=1
r+=1
title(ws,r,1,2,"2P (2-CORE) NIC SETTINGS",bg=HDR_BLUE,sz=11); r+=1
for k,v in [
    ("Queues (combined)", f"{NUM_CORES}  (ethtool -L eth2 combined {NUM_CORES})"),
    ("iperf cores", f"{NUM_CORES} iperf3 procs pinned to cores {CORES[0]}-{CORES[-1]} (one per core)"),
    ("IRQ siblings", f"cores map to SMT siblings {SIBLINGS[0]}-{SIBLINGS[-1]} (core+256)"),
    ("IRQ interleave", f"64 eth2 IRQs round-robin across the {NUM_CORES} siblings: "
                       f"IRQ i -> sibling[i % {NUM_CORES}]  (32 IRQs per sibling)"),
    ("Ring size","2048 / 2048"),("Flow control","rx OFF, tx OFF"),
    ("CQE_COMPRESSION","AGGRESSIVE(1)"),("PCI_WR_ORDERING","force_relax(1)"),
    ("irqbalance","stopped")]:
    kv(ws,r,k,v); r+=1
r+=1
title(ws,r,1,2,"TEST METHOD",bg=HDR_BLUE,sz=11); r+=1
for k,v in [("Tool",f"iperf3 — {NUM_CORES} concurrent processes, one per core; -P streams per process"),
    ("Direction","SUT clients -> LoadGen servers (unidirectional TCP Rx)"),
    ("Duration / Iterations","60 s per run, 10 iterations per -P (aggregate avg reported)"),
    ("-P sweep",f"1, 2, 4, 8, 12, 16, 20, 24, 28, 32  (same -P on all {NUM_CORES} procs)"),
    ("Throughput",f"aggregate = sum of {NUM_CORES} procs; per-core detail retained"),
    ("SDCI State","ENABLED")]:
    kv(ws,r,k,v); r+=1
r+=1
title(ws,r,1,2,"iPERF3 COMMANDS ISSUED",bg=HDR_BLUE,sz=11); r+=1
def cmd(r,label,c):
    a=ws.cell(r,1,label); b=ws.cell(r,2,c)
    a.font=Font(bold=True,color="333333"); a.alignment=Alignment(vertical="center",indent=1)
    a.fill=PatternFill("solid",fgColor=LT_GREY); a.border=BORDER
    b.font=Font(name="Consolas",size=9); b.alignment=Alignment(vertical="center",indent=1,wrap_text=True); b.border=BORDER
    ws.row_dimensions[r].height=30; return r+1
ports=f"{BASE_PORT}..{BASE_PORT+NUM_CORES-1}"
r=cmd(r,f"LoadGen  {NUM_CORES} servers", f"for p in {ports}: taskset -c <core> iperf3 -s -B 192.168.10.3 -p <p>")
r=cmd(r,f"SUT  {NUM_CORES} clients (concurrent)",
      f"taskset -c <core> iperf3 -c 192.168.10.3 -p <port> -P <N> -t 60 -J   "
      f"(core={CORES[0]}..{CORES[-1]}, port={ports})")
r=cmd(r,"  -P <N> sweep","1, 2, 4, 8, 12, 16, 20, 24, 28, 32   (10 iterations each)")
r=cmd(r,f"  force {NUM_CORES}Q combined", f"ethtool -L eth2 combined {NUM_CORES}")
r=cmd(r,"  interleave 64 IRQs",
      f"BDF=$(ethtool -i eth2|awk '/bus-info:/{{print $2}}'); SIB=({' '.join(map(str,SIBLINGS))}); i=0; "
      f"for irq in $(grep $BDF /proc/interrupts|sed -E 's/^ *([0-9]+):.*/\\1/'); do "
      f"echo ${{SIB[$((i%{NUM_CORES}))]}} > /proc/irq/$irq/smp_affinity_list; i=$((i+1)); done")

# ── Tab 2: 2-Core Study ──
ws2=wb.create_sheet("2-Core Study"); ws2.sheet_view.showGridLines=False
avgs=load_avgs(RESULT_DIR); pc=load_percore(RESULT_DIR)
plist=sorted(avgs.keys())
ncol=6+NUM_CORES
for j,w in enumerate([8,15,11,11,9,14]+[9]*NUM_CORES, start=1):
    ws2.column_dimensions[openpyxl.utils.get_column_letter(j)].width=w

title(ws2,1,1,ncol,"CX7 2-Core iPerf Core-Scaling — SDCI ENABLED",bg=AMD_RED,sz=14,ht=30)
title(ws2,2,1,ncol,
      f"eth2 (CX7 400G) | {NUM_CORES}Q | iperf cores {CORES[0]}-{CORES[-1]}, IRQs interleaved to siblings "
      f"{SIBLINGS[0]}-{SIBLINGS[-1]} | 60s x 10 iters | BIOS G2 | Venice B0",
      bg=AMD_GREY,sz=9,ht=20)

title(ws2,4,1,ncol,"2-CORE STUDY — Aggregate throughput + per-core breakdown (SDCI ENABLED)",bg=HDR_BLUE,sz=11)
hr=5
heads=["-P","Agg (Gbps)","Min","Max","CV %","Retrans"]+[f"core{c}" for c in CORES]
for j,h in enumerate(heads,start=1):
    c=ws2.cell(hr,j,h); c.font=Font(bold=True,color=WHITE); c.fill=PatternFill("solid",fgColor=HDR_BLUE)
    c.alignment=Alignment(horizontal="center",vertical="center"); c.border=BORDER
ws2.row_dimensions[hr].height=18
best=max(plist,key=lambda P:float(avgs[P][1])) if plist else None
rr=hr+1
for P in plist:
    a=avgs[P]; row=[P,float(a[1]),float(a[2]),float(a[3]),float(a[4]),float(a[5])]
    row+= [round(x,2) for x in pc.get(P,[0]*NUM_CORES)]
    for j,v in enumerate(row,start=1):
        c=ws2.cell(rr,j,v); c.alignment=Alignment(horizontal="center",vertical="center"); c.border=BORDER
        if j in (2,3,4,5) or j>6: c.number_format="0.00"
        if P==best: c.fill=PatternFill("solid",fgColor=BEST_GRN); c.font=Font(bold=True)
        elif rr%2==0: c.fill=PatternFill("solid",fgColor=LT_GREY)
    ws2.row_dimensions[rr].height=15; rr+=1

rr+=1
title(ws2,rr,1,ncol,"KEY OBSERVATIONS",bg=HDR_BLUE,sz=10); rr+=1
if plist:
    notes=[
        f"Best aggregate: -P {best} -> {float(avgs[best][1]):.2f} Gbps (CV {float(avgs[best][4]):.2f}%).",
        f"{NUM_CORES} iperf3 procs on cores {CORES[0]}-{CORES[-1]}; {NUM_CORES}Q; 64 IRQs interleaved across siblings "
        f"{SIBLINGS[0]}-{SIBLINGS[-1]} via i%{NUM_CORES} (32 IRQs per sibling).",
        f"Per-core columns show load balance across the {NUM_CORES} cores.",
        "SDCI ENABLED. Compare against 1P (~107 Gbps), 8P (~330 Gbps), 16P (~377 Gbps) to see scaling.",
    ]
    for n in notes:
        ws2.merge_cells(start_row=rr,start_column=1,end_row=rr,end_column=ncol)
        c=ws2.cell(rr,1,"•  "+n); c.alignment=Alignment(vertical="center",indent=1,wrap_text=True)
        ws2.row_dimensions[rr].height=26; rr+=1

wb.save(OUT)
print("WROTE:",OUT)
print("from:",RESULT_DIR,"| rows:",len(plist),"| best -P:",best,"->",avgs[best][1] if best else "n/a")
